# -*- coding: utf-8 -*-
"""
Build data/AirCanada_Process_Catalog.xlsx

Tabs:
  Index          cover sheet with counts and how to read the evidence tiers
  Master         one row per process, all 284
  L4 Steps       one row per L4 step across every documented process
  Systems        the system registry with evidence tiers
  EA Diagrams    the 28 architecture diagrams
  <Domain> x12   one tab per L1 domain, its processes
"""
import json, os, sys, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from taxonomy import DOMAINS, all_processes
import ac_systems

AC_RED = "D2001F"
AC_BLACK = "1A1A1A"
AC_LIGHT = "FFF0F2"

HDR_FILL = PatternFill("solid", fgColor=AC_BLACK)
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill("solid", fgColor=AC_LIGHT)
TITLE_FONT = Font(color=AC_RED, bold=True, size=16)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIER_FILL = {"A": PatternFill("solid", fgColor="D6F0E4"),
             "B": PatternFill("solid", fgColor="FDF0D5"),
             "C": PatternFill("solid", fgColor="EDEDED"),
             "U": PatternFill("solid", fgColor="FFD6DC")}


def sheet(wb, title, headers, rows, widths, table_name=None, tier_col=None):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for i, r in enumerate(rows, start=2):
        ws.append(r)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(size=9)
            if i % 2 == 0:
                cell.fill = ALT_FILL
        if tier_col:
            tv = str(ws.cell(row=i, column=tier_col).value or "")
            if tv in TIER_FILL:
                tc = ws.cell(row=i, column=tier_col)
                tc.fill = TIER_FILL[tv]
                tc.alignment = Alignment(horizontal="center", vertical="top")
                tc.font = Font(size=9, bold=True)

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    if rows and table_name:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        try:
            ws.add_table(t)
        except Exception:
            pass
    return ws


def main():
    procs = all_processes()
    pj = os.path.join(ROOT, "data", "processes.json")
    ej = os.path.join(ROOT, "data", "ea.json")
    content = json.load(open(pj, encoding="utf-8")) if os.path.exists(pj) else {}
    ea = json.load(open(ej, encoding="utf-8")).get("diagrams", []) if os.path.exists(ej) else []

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Index ----------
    ws = wb.create_sheet("Index")
    ws["A1"] = "Air Canada Process Catalog"; ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.date.today().isoformat()}"
    ws["A2"].font = Font(size=9, color="6B6B6B")
    ws["A4"] = "Contents"; ws["A4"].font = Font(bold=True, size=12, color=AC_RED)

    documented = sum(1 for p in procs if p["pid"] in content)
    steps_total = sum(len(c.get("l4_steps", [])) for c in content.values())
    rows = [
        ("Master", "One row per process", len(procs)),
        ("L4 Steps", "One row per documented L4 step", steps_total),
        ("Systems", "System registry with evidence tiers", len(ac_systems.SYSTEMS)),
        ("EA Diagrams", "Domain landscapes and cross-domain flows", len(ea)),
    ] + [(d["name"][:31], f'{d["code"]} domain processes',
          sum(len(g["p"]) for g in d["l2"])) for d in DOMAINS]

    ws.append([]); ws.append(["Tab", "Contents", "Rows"])
    for c in range(1, 4):
        cell = ws.cell(row=6, column=c); cell.fill = HDR_FILL; cell.font = HDR_FONT
    for r in rows:
        ws.append(list(r))
    for i in range(7, 7 + len(rows)):
        for c in range(1, 4):
            ws.cell(row=i, column=c).font = Font(size=9)
            ws.cell(row=i, column=c).border = BORDER

    base = 7 + len(rows) + 2
    ws.cell(row=base, column=1, value="Coverage").font = Font(bold=True, size=12, color=AC_RED)
    for j, (k, v) in enumerate([
        ("L1 domains", len(DOMAINS)),
        ("L2 groups", sum(len(d["l2"]) for d in DOMAINS)),
        ("L3 processes", len(procs)),
        ("Processes documented", documented),
        ("Processes queued", len(procs) - documented),
        ("L4 steps documented", steps_total),
    ], start=1):
        ws.cell(row=base + j, column=1, value=k).font = Font(size=9)
        ws.cell(row=base + j, column=2, value=v).font = Font(size=9, bold=True)

    base2 = base + 9
    ws.cell(row=base2, column=1, value="Evidence tiers").font = Font(bold=True, size=12, color=AC_RED)
    for j, t in enumerate("ABCU", start=1):
        ws.cell(row=base2 + j, column=1, value=t).fill = TIER_FILL[t]
        ws.cell(row=base2 + j, column=1).font = Font(size=9, bold=True)
        ws.cell(row=base2 + j, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=base2 + j, column=2, value=ac_systems.TIER_LABEL[t]).font = Font(size=9)

    warn = base2 + 6
    ws.cell(row=warn, column=1,
            value="Tier C is an industry pattern and tier U is an open question. Neither is a claim of "
                  "fact about Air Canada and neither should be asserted in a bid without confirmation.")
    ws.cell(row=warn, column=1).font = Font(size=9, italic=True, color=AC_RED)
    for c, w in zip("ABC", (34, 62, 12)):
        ws.column_dimensions[c].width = w

    # ---------- Master ----------
    master = []
    for p in procs:
        c = content.get(p["pid"], {})
        master.append([
            p["pid"], p["l1_code"], p["l1_name"], p["l2_code"], p["l2_name"], p["l3_name"],
            "Documented" if c else "Queued",
            c.get("description", ""), c.get("trigger", ""), c.get("outcome", ""),
            c.get("ac_notes", ""),
            len(c.get("l4_steps", [])),
            " | ".join(c.get("systems", [])),
            " | ".join(c.get("kpis", [])),
            " | ".join(c.get("risks", [])),
            f'{p["l1_slug"]}/{p["l2_slug"]}/{p["pid"].lower()}/',
        ])
    sheet(wb, "Master",
          ["PID", "L1", "L1 Domain", "L2", "L2 Group", "L3 Process", "Status", "Description",
           "Trigger", "Outcome", "Air Canada notes", "Steps", "Systems", "KPIs", "Risks", "URL path"],
          master,
          [15, 6, 26, 6, 26, 42, 12, 60, 42, 42, 62, 7, 44, 52, 58, 46],
          table_name="MasterTbl")

    # ---------- L4 Steps ----------
    steps = []
    by_pid = {p["pid"]: p for p in procs}
    for pid, c in content.items():
        p = by_pid[pid]
        phases = c.get("phases", [])
        for s in c.get("l4_steps", []):
            ph = int(str(s.get("step", "1")).split(".")[0])
            steps.append([
                pid, p["l1_code"], p["l3_name"],
                s.get("step"), phases[ph - 1] if ph <= len(phases) else f"Phase {ph}",
                s.get("name"), s.get("role"), s.get("system"),
                ac_systems.resolve(s.get("system", ""))[3],
                s.get("input"), s.get("output"), s.get("kpi"),
                s.get("decision_point"), s.get("exception"), s.get("pain_point"),
            ])
    sheet(wb, "L4 Steps",
          ["PID", "L1", "Process", "Step", "Phase", "Activity", "Role", "System", "Tier",
           "Input", "Output", "KPI", "Decision", "Exception", "Pain point"],
          steps,
          [15, 6, 40, 7, 28, 40, 28, 30, 6, 34, 34, 42, 9, 10, 54],
          table_name="StepsTbl", tier_col=9)

    # ---------- Systems ----------
    sysrows = []
    for name in ac_systems.all_systems():
        cls, tip, tier, vendor = ac_systems.SYSTEMS[name]
        used = sorted({pid for pid, c in content.items() if name in c.get("systems", [])})
        sysrows.append([name, vendor, tier, ac_systems.TIER_LABEL[tier], tip,
                        len(used), " ".join(used)])
    sheet(wb, "Systems",
          ["System", "Vendor", "Tier", "Tier meaning", "Description",
           "Processes using", "PIDs"],
          sysrows, [38, 22, 6, 46, 74, 14, 50],
          table_name="SystemsTbl", tier_col=3)

    # ---------- EA Diagrams ----------
    earows = [[d["id"], d["title"], d["kind"], d["blurb"],
               " | ".join(d.get("systems", [])), " | ".join(d.get("notes", [])),
               f'ea-diagrams/{d["slug"]}/'] for d in ea]
    sheet(wb, "EA Diagrams",
          ["ID", "Title", "Kind", "Description", "Systems", "Reading notes", "URL path"],
          earows, [12, 48, 20, 66, 52, 80, 36], table_name="EATbl")

    # ---------- Per-domain tabs ----------
    for d in DOMAINS:
        rows_d = []
        for p in [x for x in procs if x["l1_code"] == d["code"]]:
            c = content.get(p["pid"], {})
            rows_d.append([p["pid"], p["l2_code"], p["l2_name"], p["l3_name"],
                           "Documented" if c else "Queued",
                           len(c.get("l4_steps", [])),
                           c.get("description", ""),
                           " | ".join(c.get("systems", [])),
                           " | ".join(c.get("kpis", []))])
        sheet(wb, d["name"],
              ["PID", "L2", "L2 Group", "L3 Process", "Status", "Steps",
               "Description", "Systems", "KPIs"],
              rows_d, [15, 6, 28, 44, 12, 7, 62, 44, 52],
              table_name=f'Tbl{d["code"]}')

    out = os.path.join(ROOT, "data", "AirCanada_Process_Catalog.xlsx")
    wb.save(out)
    print(f"Wrote {out} ({os.path.getsize(out)/1024:.0f} KB)")
    print(f"  {len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames[:6])} ...")
    print(f"  Master {len(master)} rows | L4 Steps {len(steps)} rows | "
          f"Systems {len(sysrows)} rows | EA {len(earows)} rows")


if __name__ == "__main__":
    main()
